from pathlib import Path

from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet


ROOT_FOLDER = Path(__file__).parent
WORKBOOK_PATH = ROOT_FOLDER/ 'workbook.xlsx'

workbook = Workbook()
worksheet: Worksheet = workbook.active

worksheet.cell(1, 1, 'Nome')
worksheet.cell(1, 2, 'Idade')
worksheet.cell(1, 3, 'Nota')

students = [
    ['Felipe', 26, 10],
    ['Livia', 24, 10],
    ['Nemo', 1, 10],
    ['Dory', 0.3, 10]
]

for i, student_row in enumerate(students, start=1):
    print(i, student_row)
workbook.save(WORKBOOK_PATH)